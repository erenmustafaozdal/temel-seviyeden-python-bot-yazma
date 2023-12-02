from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import logging
import sys
import os
import urllib.request
from openpyxl import Workbook, load_workbook
from datetime import datetime


# görselleri indireceğimiz klasör yoksa oluşturalım
images_dir = "./images"
if not os.path.exists(images_dir):
    os.mkdir(images_dir)


# excel dosyası yoksa oluştur ve sütun başlıklarını yaz, varsa aç
excel_path = "amazon_products.xlsx"
if not os.path.exists(excel_path):
    wb = Workbook()
    ws = wb.active
    # dosya oluşturuldu ve başlıkları yazılacak
    ws.append(["TARİH", "ID", "AD", "GÖRSEL", "FİYAT"])
else:
    wb = load_workbook(excel_path)
    ws = wb.active

komut_log = logging.StreamHandler(stream=sys.stdout)

logging.basicConfig(
    handlers=[komut_log],
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    datefmt="%d-%m-%Y %H:%M:%S"
)

logger = logging.getLogger(__name__)

# Tarayıcı nesnesini oluşturalım
chrome_install = ChromeDriverManager().install()
chrome_service = Service(chrome_install)
driver = webdriver.Chrome(service=chrome_service)
driver.maximize_window()

# amazon sayfasına gidelim
driver.get("https://amazon.com.tr")

# sayfanın yüklendiğinde emin olalım (arama kutucuğu sayfada görünene kadar bekle)
while True:
    try:
        sleep(5)
        driver.find_element(By.ID, "twotabsearchtextbox")
        break
    except:
        driver.refresh()

# eğer varsa çerezleri kabul et
try:
    driver.find_element(By.ID, "sp-cc-accept").click()
except:
    logger.error("Çerezleri kabul et tuşu bulunamadı.")

# arama yap
driver.minimize_window()
search = "laptop sırt çantası"  # input("Ne aramak isterseniz? ")
driver.maximize_window()
driver.find_element(By.ID, "twotabsearchtextbox").send_keys(
    search + Keys.ENTER)

# ürünleri gez
products = driver.find_elements(
    By.XPATH, "//div[@data-component-type='s-search-result']")
for product in products:
    # ürünün id'sini alalım
    id = product.get_attribute("data-asin")
    logger.info(f"{id} ID'li ürün bilgileri alınıyor...")

    # ürünün adını alalım
    name = product.find_element(By.TAG_NAME, "h2").text
    logger.info(f"----- Ürün adı: {name}")

    # ürünün resim linkini (<img src="">) alalım
    src = product.find_element(By.TAG_NAME, "img").get_attribute("src")
    image_path = f"{images_dir}/{id}.jpg"
    urllib.request.urlretrieve(src, image_path)
    logger.info(f"----- Ürün resim linki: {src} (indirildi)")

    # ürünün fiyatını alalım
    try:
        lira = product.find_element(By.CLASS_NAME, "a-price-whole").text
        kurus = product.find_element(By.CLASS_NAME, "a-price-fraction").text
        price = float(f"{lira}.{kurus}")
        logger.info(f"----- Ürünün fiyatı: {price}")
    except:
        price = ""
        logger.error("----- Ürün fiyatı bulunamadı.")

    # Excel kayıt işlemleri
    ws.append([
        datetime.now(),
        id,
        name,
        f'=HYPERLINK("{os.getcwd()}/{image_path}", "GÖRSEL")',
        price
    ])


wb.save(excel_path)
