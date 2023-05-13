from openpyxl import Workbook, load_workbook
import os

excel_dosya = "./ogrenciler.xlsx"

# Yeni bir excel dosyası oluşturalım
if not os.path.exists(excel_dosya):
    calisma_kitabi = Workbook()

    # Aktif çalışma sayfasını açalım
    cs = calisma_kitabi.active

    # Verileri yazma
    cs["A1"] = "Adı"
    cs["B1"] = "Soyadı"
    cs["C1"] = "Yaşı"
else:
    calisma_kitabi = load_workbook(excel_dosya)
    cs = calisma_kitabi.active

# Yeni satırlar ekleyelim
cs.append(["Eren", "Özdal", 37])
cs.append(["Ahmet", "Yılmaz", 22])

# çalışma kitabını kaydet
calisma_kitabi.save(excel_dosya)
