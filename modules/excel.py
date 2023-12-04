import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class Excel:
    """
    Excel çalışma kitabınını veritabanı gibi temsil eden sınıf
    """

    def __init__(self, file_path: str = None) -> None:
        """
        Excel sınıfının yapıcı metodu

        Args:
            file_path (str, optional): Excel veritabanı dosyasının adı. Defaults to None.
        """
        self.file_path = file_path or "./db/db.xlsx"

        # klasör yolu alınır
        directory = os.path.dirname(file_path)
        # eğer klasör yoksa oluşturulur
        if not os.path.exists(directory):
            os.makedirs(directory)

        # çalışma kitabı alınır
        self.wb: Workbook = self.get_workbook(file_path)

    def get_workbook(self, file_path: str) -> Workbook:
        """
        Excel dosyasını yükler veya yeni bir çalışma kitabı oluşturur.

        Args:
            file_path (str): Excel dosyasının adı.

        Returns:
            Workbook: Yüklenen veya oluşturulan çalışma kitabı.
        """
        # eğer dosya varsa yükle döndür
        if os.path.exists(file_path):
            return load_workbook(file_path)

        # eğer dosya yoksa oluştur ve döndür
        return Workbook()

    def save(self):
        # Dosyayı kaydet
        self.wb.save(self.file_path)


class ExcelSheet:
    """
    Excel çalışma sayfasını veritabanı tablosu gibi temsil eden sınıf
    """

    sheet_name = ""
    column_names = []

    def __init__(self, wb: Workbook) -> None:
        self.wb = wb
        # Eğer çalışma sayfası yoksa yeni bir tane oluştur
        self.sheet = self.get_worksheet(self.sheet_name)
        # tablo sütun başlıkları oluşturulur
        self.create_column_names(self.column_names)

    def get_worksheet(self, sheet_name: str) -> Worksheet:
        # eğer çalışma sayfası varsa döndür
        if sheet_name in self.wb.sheetnames:
            return self.wb[sheet_name]

        # eğer varsayılan çalışma sayfası adı varsa
        if "Sheet" in self.wb.sheetnames:
            sheet = self.wb["Sheet"]
            sheet.title = sheet_name
            return sheet

        # eğer hiçbiri yoksa yeni oluştur
        return self.wb.create_sheet(sheet_name)

    def create_column_names(self, column_names):
        # Sütun başlıklarını ekle
        for i, column in enumerate(column_names, start=1):
            self.sheet.cell(1, i, column)

    def insert(self, data: dict):
        # Veriyi Excel sheet'ine eklemek için kullanılır.
        row = [data.get(column, None) for column in self.column_names]
        self.sheet.append(row)
